# -*-coding=utf-8-*-

import pandas as pd
from openpyxl import load_workbook
import copy
import _sqlite3
from os import path, getcwd

supply_dict = None
demand_dict = None
arcs_list = None


class Node(object):
    def __init__(self, code, name, node_type, volume=0, province=''):
        self.code = code
        self.name = name
        self.type = node_type
        self.volume = volume
        self.province = province
        self.tra_cost = 0
        self.outlines = []
        self.in_degree = 0
        self.deepth = 0
        self.up_arcs = []
        self.sup_vol_dict = {name: volume} if node_type == 'supply' else {}
        self.sup_rat_dict = {name: 1.0} if node_type == 'supply' else {}


class Line(object):
    def __init__(self, code, name, up_node: Node, down_node: Node, fee, mileage, volume):
        self.code = code
        self.name = name
        self.up_node = up_node
        self.down_node = down_node
        self.fee = fee
        self.mileage = mileage
        self.volume = volume


# def get_node_dict(file_path, sheet_name, node_type):
#     node_dict = {}
#     df = pd.read_excel(file_path, sheet_name=sheet_name)
#     for index, row in df.iterrows():
#         node = Node(row['code'], row['name'], node_type, row['volume'] if node_type == 'supply' else 0)
#         node_dict[row['code']] = node
#     return (df, node_dict) if node_type == 'demand' else node_dict
#
#
# def get_arcs_list(file_path, sheet_name, node_dict):
#     arcs_list = []
#     df = pd.read_excel(file_path, sheet_name=sheet_name)
#     for index, row in df.iterrows():
#         if row['volume'] == 0:
#             continue
#         up_code = row['point_code_up']
#         down_code = row['point_code_down']
#         line = Line(row['code'], row['name'], node_dict[up_code], node_dict[down_code],
#                     row['price'], row['mileage'], row['volume'])
#         arcs_list.append(line)
#     return arcs_list


def evenly_split_process():
    global supply_dict, arcs_list

    # 将气源节点添加到链表中
    linklist = []
    for supply in supply_dict.values():
        # if supply.volume == 0: continue
        linklist.append(supply)

    while len(linklist):
        node = linklist.pop(0)
        for arc in node.outlines:
            down_node = arc.down_node
            down_node.volume += arc.volume
            down_node.tra_cost += arc.volume / node.volume * node.tra_cost if node.volume != 0 else 0
            down_node.tra_cost += arc.volume * arc.fee * arc.mileage
            for supply_name, supply_ratio in node.sup_rat_dict.items():
                volume_add = arc.volume * supply_ratio
                volume_update = down_node.sup_vol_dict.get(supply_name, 0) + volume_add
                down_node.sup_vol_dict[supply_name] = volume_update
            down_node.in_degree -= 1
            if down_node.in_degree == 0:
                for supply_name, supply_volume in copy.deepcopy(down_node.sup_vol_dict).items():
                    if supply_volume == 0:
                        del down_node.sup_vol_dict[supply_name]
                        continue
                    ratio = supply_volume / down_node.volume if down_node.volume != 0 else 0
                    down_node.sup_rat_dict[supply_name] = ratio
                linklist.append(down_node)


def output(file_path):
    total = 0
    global demand_dict
    pd.set_option('max_colwidth', 200)
    result_df = pd.DataFrame(columns=('code', 'name', 'province', 'volume', 'tra_cost', 'sup_ratio', 'sup_vol'))
    for index, node in enumerate(demand_dict.values()):
        result_df.loc[index] = [node.code, node.name, node.province, node.volume, node.tra_cost,
                                percentage_trans(node.sup_rat_dict), node.sup_vol_dict]
    result_df.to_excel(file_path, sheet_name='result_cus')
    return result_df


def demand_group(file_path):
    global demand_dict
    pd.set_option('max_colwidth', 200)
    result_df = pd.DataFrame(columns=('province', 'supply', 'volume'))
    for node in demand_dict.values():
        if node.province == '': continue
        for supply, volume in node.sup_vol_dict.items():
            result_df.loc[result_df.shape[0]] = [node.province, supply, volume]

    result_df1 = result_df.groupby(['province', 'supply']).sum()
    sub_df = result_df1.groupby('province').sum()
    result_df1['ratio'] = result_df1.div(sub_df, axis=0)['volume']
    result_df1 = result_df1.sort_values(by=['province', 'ratio'], ascending=[True, False])
    result_df1['ratio'] = result_df1['ratio'].apply(lambda x: '%.2f%%' % (x * 100))
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    result_df1.to_excel(writer, sheet_name='result_prov')
    writer.save()
    # print(result_df1)

    result_df2 = result_df.groupby(['supply', 'province']).sum()
    sub_df = result_df2.groupby('supply').sum()
    result_df2['ratio'] = result_df2.div(sub_df, axis=0)
    result_df2 = result_df2.sort_values(by=['supply', 'ratio'], ascending=[True, False])
    result_df2['ratio'] = result_df2['ratio'].apply(lambda x: '%.2f%%' % (x * 100))
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    result_df2.to_excel(writer, sheet_name='result_supp')
    writer.save()
    # print(result_df2)


def percentage_trans(ratio_dict):
    for key, ratio in ratio_dict.items():
        ratio_dict[key] = '%.2f%%' % (ratio * 100)
    return ratio_dict


def read_sqlite3(file_path, year_id):
    global supply_dict, demand_dict, arcs_list
    year_id = str(year_id)
    pd.set_option('max_colwidth', 200)
    with _sqlite3.connect(file_path) as con:
        # 获取场站节点信息
        select_sql = 'SELECT NodeID id, Caption name FROM tbl_Input_Node_Static'
        station_df = pd.read_sql_query(select_sql, con)
        station_dict = {}
        for row in station_df.itertuples():
            node = Node(row.id, row.name, 'station')
            station_dict[node.code] = node

        # 获取管段信息
        select_sql = 'SELECT a.PipeID id, a.Caption name, a.UpNodeID up_node_id, b.Length mileage, ' \
                     'a.DownNodeID down_node_id, b.Length length, c.YearUnitAlterableCost price, ' \
                     'c.YearUpFlowRate volume ' \
                     'FROM tbl_Input_Pipe_Static a, tbl_Input_Pipe_Process_Fixed b, tbl_Output_Pipe_Year c ' \
                     'ON a.PipeID = c.PipeID and a.PipeID = b.PipeID ' \
                     'WHERE c.CaseID = 1 AND c.YearID = ' + year_id
        arcs_df = pd.read_sql_query(select_sql, con)
        arcs_list = []
        for row in arcs_df.itertuples():
            if row.volume == 0:
                continue
            elif row.volume > 0:
                line = Line('P' + str(len(arcs_list)), row.name, station_dict[row.up_node_id],
                            station_dict[row.down_node_id], row.price, row.mileage, row.volume)
            else:
                line = Line('P' + str(len(arcs_list)), row.name, station_dict[row.down_node_id],
                            station_dict[row.up_node_id], row.price, row.mileage, -row.volume)
            arcs_list.append(line)

        # 获取气源节点信息
        select_sql = 'SELECT a.SourceID id, a.Caption name, a.NodeID node_id, b.YearFlowRate volume ' \
                     'FROM tbl_Input_Source_Static a INNER JOIN tbl_Output_Source_Year b ' \
                     'ON a.SourceID = b.GasSourceID ' \
                     'WHERE b.CaseID = 1 AND b.YearID = ' + year_id
        supply_df = pd.read_sql_query(select_sql, con)
        supply_dict = {}
        for row in supply_df.itertuples():
            if row.volume == 0: continue
            node = Node('S' + str(len(supply_dict)), row.name, 'supply', row.volume)
            supply_dict[node.code] = node
            line = Line('P' + str(len(arcs_list)), row.name, node, station_dict[row.node_id], 0, 1, row.volume)
            arcs_list.append(line)

        # 获取用户节点信息
        select_sql = 'SELECT a.ClientID id, a.Caption name, a.NodeID node_id, b.YearFlowRate volume, ' \
                     'a.Province province ' \
                     'FROM tbl_Input_Client_Static a INNER JOIN tbl_Output_Client_Year b ' \
                     'ON a.ClientID = b.GasClientID ' \
                     'WHERE b.CaseID = 1 AND b.YearID = ' + year_id
        demand_df = pd.read_sql_query(select_sql, con)
        demand_dict = {}
        for row in demand_df.itertuples():
            if row.volume == 0: continue
            node = Node('L' + str(len(demand_dict)), row.name, 'demand', province=row.province)
            demand_dict[node.code] = node
            line = Line('P' + str(len(arcs_list)), row.name, station_dict[row.node_id], node, 0, 1, row.volume)
            arcs_list.append(line)

        # 获取储气库信息
        select_sql = 'SELECT a.StorageID id, a.Caption name, a.NodeID node_id, b.YearFlowRate volume ' \
                     'FROM tbl_Input_Storage_Static a INNER JOIN tbl_Output_Storage_Year b ' \
                     'ON a.StorageID = b.GasStorageID ' \
                     'WHERE b.CaseID = 1 AND b.YearID = ' + year_id
        other_df = pd.read_sql_query(select_sql, con)
        for row in other_df.itertuples():
            if row.volume == 0: continue
            elif row.volume > 0:
                node = Node('L' + str(len(demand_dict)), row.name, 'demand')
                demand_dict[node.code] = node
                line = Line('P' + str(len(arcs_list)), row.name, station_dict[row.node_id], node, 0, 1, row.volume)
                arcs_list.append(line)
            else:
                node = Node('S' + str(len(supply_dict)), row.name, 'supply', -row.volume)
                supply_dict[node.code] = node
                line = Line('P' + str(len(arcs_list)), row.name, node, station_dict[row.node_id], 0, 1, -row.volume)
                arcs_list.append(line)

        # 获取接收站信息
        select_sql = 'SELECT a.TankID id, a.Caption name, a.UpNodeID up_node_id, a.DownNodeID down_node_id, ' \
                     'b.YearUpFlowRate volume ' \
                     'FROM tbl_Input_Tank_Static a INNER JOIN tbl_Output_Tank_Year b ' \
                     'ON a.TankID = b.TankID ' \
                     'WHERE b.CaseID = 1 AND b.YearID = ' + year_id
        other_df = pd.read_sql_query(select_sql, con)
        for row in other_df.itertuples():
            if row.volume == 0: continue
            node = Node('T' + str(row.Index), row.name, 'Tank')
            line = Line('P' + str(len(arcs_list)), row.name, station_dict[row.up_node_id], node, 0, 1, row.volume)
            arcs_list.append(line)
            line = Line('P' + str(len(arcs_list)), row.name, node, station_dict[row.down_node_id], 0, 1, row.volume)
            arcs_list.append(line)

        # 获取损耗信息
        select_sql = 'SELECT a.FixedWastingGasID id, a.Caption name, a.NodeID node_id, b.YearFlowRate volume ' \
                     'FROM tbl_Input_FixedWastingGas_Static a INNER JOIN tbl_Output_FixedWastingGas_Year b ' \
                     'ON a.FixedWastingGasID = b.FixedWastingGasID ' \
                     'WHERE b.CaseID = 1 AND b.YearID = ' + year_id
        other_df = pd.read_sql_query(select_sql, con)
        for row in other_df.itertuples():
            if row.volume == 0: continue
            node = Node('L' + str(len(demand_dict)), row.name, 'demand')
            demand_dict[node.code] = node
            line = Line('P' + str(len(arcs_list)), row.name, station_dict[row.node_id], node, 0, 1, row.volume)
            arcs_list.append(line)

    return supply_dict, demand_dict, arcs_list


def accul(db_file_path, year_id):
    supply_dict, demand_dict, arcs_list = read_sqlite3(db_file_path, year_id)
    evenly_split_process(supply_dict, arcs_list)


# 初始化各节点的下游路径集合
def ini_outlines():
    global supply_dict, arcs_list
    # 初始化各节点的下游路径集合和入度值
    for arc in arcs_list:
        arc.up_node.outlines.append(arc)
        arc.down_node.in_degree += 1


# 计算气源点supply_node就近销售的用户
def sales_nearby_process(supply_node):
    # print(supply_node.name, supply_node.volume)
    global supply_dict, arcs_list
    # demandlist = []
    # 填充linklist内的元素为[节点，深度，上游管线]
    linklist = [(supply_node, 0, [])]
    # print(supply_node.code, supply_node.name, supply_node.volume)
    while len(linklist):
        node, pre_deepth, pre_up_arcs = linklist.pop(0)
        if supply_node.code == 'S8':
            print('取出一个点')
        # print('node:', supply_node.name, node.name, supply_node.volume, node.deepth)
        # print('🔺🔺', node.code, node.name)
        # for arc in node.outlines:
        # print('🔺', arc.code, arc.name)
        for arc in node.outlines:
            if supply_node.code == 'S8':
                print('⭐', arc.code, arc.name, arc.up_node.name, arc.up_node.code, arc.down_node.name,
                arc.down_node.code, arc.mileage)
            down_node = arc.down_node
            cur_deepth = pre_deepth + arc.mileage
            cur_up_arcs = pre_up_arcs[:]
            cur_up_arcs.append(arc)
            if down_node.type == 'demand':
                # demandlist.append(down_node)
                # 找一下上游管段的最小输量
                min_trans_volume = supply_node.volume
                for arc in cur_up_arcs:
                    if min_trans_volume > arc.volume:
                        min_trans_volume = arc.volume
                # print(down_node.name, '上游最小输量：', min_trans_volume, arc.volume)
                volume_add = arc.volume if arc.volume < min_trans_volume else min_trans_volume
                if supply_node.code == 'S8':
                    print(supply_node.code, supply_node.name, supply_node.volume)
                    print(down_node.code, down_node.name, pre_deepth, down_node.volume, volume_add)
                if volume_add == 0: continue
                # 计算用户承担的管输费
                tra_cost_add = 0
                for arc in cur_up_arcs:
                    tra_cost_add += volume_add * arc.fee * arc.mileage
                down_node.tra_cost += tra_cost_add
                down_node.volume += volume_add
                supply_node.volume -= volume_add
                if supply_node.name in down_node.sup_vol_dict:
                    down_node.sup_vol_dict[supply_node.name] += volume_add
                else:
                    down_node.sup_vol_dict[supply_node.name] = volume_add
                for arc in cur_up_arcs:  # 流过的路径减去相应的流量
                    arc.volume -= volume_add
                if supply_node.volume <= 1e-15: break
                # if volume_add < supply_node.volume:
                #     # down_node.volume = arc.volume
                #     down_node.sup_vol_dict[supply_node.name] = volume_add
                #     # down_node.sup_rat_dict[supply_node.name] = 1
                #     supply_node.volume -= volume_add
                #     for arc in down_node.up_arcs:  # 流过的路径减去相应的流量
                #         # print(arc.up_node.name, arc.down_node.name, arc.volume)
                #         arc.volume -= volume_add
                #         # print(arc.up_node.name, arc.down_node.name, arc.volume)
                #         if arc.volume < 0:
                #             print(supply_node.code, supply_node.name)
                #             print('ririririririririiri', arc.up_node.name, arc.volume)
                #     # print(supply_node.code, supply_node.name, supply_node.volume, down_node.code, down_node.name,
                #     #       down_node.volume, down_node.sup_vol_dict.values(), down_node.province)
                # else:  # volume_add >= supply_node.volume
                #     # down_node.volume = supply_node.volume
                #     down_node.sup_vol_dict[supply_node.name] = supply_node.volume
                #     for arc in down_node.up_arcs:  # 流过的路径减去相应的流量
                #         # print(arc.up_node.name, arc.down_node.name, arc.volume)
                #         arc.volume -= supply_node.volume
                #         # print(arc.up_node.name, arc.down_node.name, arc.volume)
                #     supply_node.volume = 0
                #     # print(supply_node.code, supply_node.name, supply_node.volume, down_node.code, down_node.name,
                #     #       down_node.volume, down_node.sup_vol_dict.values(), down_node.province)
                #     break

            else:  # 按深度大小排序，小的排在前面
                linklist.append((down_node, cur_deepth, cur_up_arcs))
                index = len(linklist) - 1
                while index > 0:
                    if linklist[index][1] < linklist[index-1][1]:
                        linklist[index], linklist[index-1] = linklist[index-1], linklist[index]
                        index -= 1
                    else: break
            if supply_node.code == 'S8':
                print(list(map(lambda x: (x[0].code, x[1]), linklist)))
                # for a in linklist:
                #     print('※※※※', a.code, a.name, a.deepth)
        if supply_node.volume <= 1e-15:
            break
    # print('--', supply_node.code, supply_node.name, supply_node.volume)


# 计算气源点supply_list就近销售的用户
def sales_nearby_supply(supply_list):
    # 给定就近销售的资源名称
    nearby_name_list = ['大庆油田其余', '辽河油田其余', '吉林油田其余', '新疆油田其余', '大港油田其余', '华北油田其余',
                        '吐哈油田其余', '冀东油田其余', '煤层气公司致密气其余', '海南福山油田其余', '沁水煤层气其余',
                        '煤层气公司煤层气', '浙江煤层气其余', '新疆庆华', '大唐克旗煤制气', '浙江页岩气',
                        '买断华北煤层气(华油买断山西地方煤制气)', '中海油天津LNG通过滨达管道大港互联', '中石化天津LNG南港互联',
                        '中石化天津LNG宝坻互联', '中石化安济线安平互联', '中石化天津LNG沧州互联', '中石化鄂安沧线安平互联',
                        '中海油蒙西煤制气文安互联', '中石化鄂安沧线兴县互联', '中海油蒙西煤制气应县互联',
                        '中石化榆济线临汾互联', '阜新煤制气沈阳互联', '川气东送和上海LNG通过上海管网白鹤互联',
                        '川气东送和上海LNG通过上海管网金山互联', '中石化川气东送通过江苏省管网无锡互联',
                        '中石化川气东送青山互联', '港华储气库金坛互联', '川气东送和宁波LNG通过浙江管网长兴互联',
                        '川气东送和宁波LNG通过浙江管网萧山互联', '川气东送通过安徽管网合肥互联', '川气东送通过安徽管网芜湖互联',
                        '中海油滨海LNG滁州互联', '中海油福建LNG漳州互联', '中海油福建LNG福州互联',
                        '中石化川气东送通过江西管网新余互联', '中石化川气东送通过江西管网鹰潭互联',
                        '中石化川气东送通过江西管网九江互联', '中石化青岛LNG高密互联', '中石化榆济线齐河互联',
                        '中海油莱威线烟台互联', '中石化新粤浙三门峡互联', '中石化川气东送武穴互联',
                        '中石化川气东送通过武汉高压管网黄陂互联', '中石化川气东送仙桃互联', '中石化新粤浙枣阳互联',
                        '中海油LNG通过广东管网鳌头首站互联', '中海油LNG通过广东管网广洲互联', '中海油大鹏LNG清溪互联',
                        '中海油揭阳LNG揭阳互联', '中石化广西管道钦州互联', '中石化广西管道来宾互联', '中石化桂渝管道都匀互联',
                        '武汉高压管网军山互联', '川气东送通过池庐线和合肥环网合肥互联']
    for supply in supply_list:
        print(supply.code, supply.name, supply.volume)
    # 按资源量由小到大的顺序进行就近销售
    supply_list.sort(key=lambda x: x.volume, reverse=False)  # 按资源量大小从小到大排序
    for supply in supply_list[:]:
        sales_nearby_process(supply)

    print()
    for index, supply in enumerate(supply_list):
        print(index, supply.code, supply.name, supply.volume)


def process_per_year(db_file_path, year):
    print('正在计算{}年的方案：'.format(year))
    read_sqlite3(db_file_path, year - 2012)
    report()
    ini_outlines()
    nearby_list = list(supply_dict.values())
    sales_nearby_supply(nearby_list)
    # evenly_split_process()
    report()
    output('E:/工作/规划院/20201027资源标签化/模型测试文件夹/检查精度问题2{}.xlsx'.format(year))
    demand_group('E:/工作/规划院/20201027资源标签化/模型测试文件夹/检查精度问题2{}.xlsx'.format(year))
    print('计算完成{}年的方案。'.format(year))


def report():
    tra_total = 0
    for key, value in demand_dict.items():
        tra_total += value.tra_cost
    print('用户tra_total:', tra_total)
    tra_total = 0
    for arc in arcs_list:
        tra_total += arc.volume * arc.mileage * arc.fee
    print('管输tra_total:', tra_total)

    total_volume = 0
    for demand in demand_dict.values():
        total_volume += demand.volume
    print('demand_volume_total:', total_volume)
    total_volume = 0
    for supply in supply_dict.values():
        total_volume += supply.volume
    print('supply_volume_total:', total_volume)


if __name__ == '__main__':
    # read_sqlite3('E:/工作/规划院/20201027资源标签化/20200408.db', 2020 - 2012)
    # ini_outlines()

    # nearby_list = list(supply_dict.values())
    # sales_nearby_supply(nearby_list)
    # nearby_list.sort(key=lambda x: x.volume, reverse=False)  # 按资源量大小从小到大排序
    # evenly_split_process()
    # output('E:/工作/规划院/20201027资源标签化/测算资源构成/测试20201.xlsx')
    # demand_group('E:/工作/规划院/20201027资源标签化/测算资源构成/测试20201.xlsx')

    db_file_path = 'E:/工作/规划院/20201027资源标签化/20200408.db'
    case = [2025]
    for year in case:
        process_per_year(db_file_path, year)
